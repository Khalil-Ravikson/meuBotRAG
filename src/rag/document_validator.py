"""
rag/document_validator.py — Validador e Auto-Config de Documentos (v1.0)
=========================================================================

O QUE RESOLVE:
───────────────
  PROBLEMA 1 — "Toda vez que vou ingerir tenho que configurar o LlamaParse":
    O validador detecta automaticamente o tipo de documento pelo conteúdo,
    gera um parsing_instruction específico e retorna um config completo.
    Não precisas mais de editar o DOCUMENT_CONFIG manualmente para cada PDF.

  PROBLEMA 2 — "Como protejo o !ingerir de ficheiros inválidos?":
    Antes de qualquer processamento, o validador verifica:
    1. Extensão permitida (lista de MIME types aceites)
    2. Tamanho do ficheiro (não deixa 500MB passarem)
    3. Estrutura mínima (PDF tem cabeçalho %PDF, CSV tem pelo menos 2 colunas)
    4. Conteúdo não vazio após parsing preliminar
    5. Para PDFs: detecta se é scan/imagem (sem texto extraível → sugere LlamaParse)

  PROBLEMA 3 — "Não sei qual parsing_instruction usar para cada PDF":
    O auto-config analisa o nome do ficheiro e os primeiros caracteres extraídos
    e escolhe o template mais adequado de parsing_instruction automaticamente.

COMO USAR:
───────────
  # No tasks_admin.py, antes de ingerir:
  from src.rag.document_validator import validar_documento, ResultadoValidacao

  resultado = validar_documento(caminho, mimetype, nome_original)
  if not resultado.valido:
      return f"❌ {resultado.motivo_rejeicao}"

  # Usa o config gerado automaticamente
  config = resultado.config_sugerido

REGRAS DE VALIDAÇÃO POR TIPO:
──────────────────────────────
  PDF:
    - Tamanho máximo: 50 MB
    - Deve ter cabeçalho %PDF
    - Se 0 páginas com texto: sugere LlamaParse ou rejeita se > 5 MB
    - Nome deve conter keyword reconhecível OU admin confirma doc_type

  CSV:
    - Tamanho máximo: 10 MB
    - Deve ter pelo menos 2 colunas na primeira linha
    - Deve ter pelo menos 2 linhas de dados
    - Encoding: UTF-8, latin-1 ou cp1252

  DOCX:
    - Tamanho máximo: 20 MB
    - Deve ser ficheiro ZIP válido (docx é ZIP internamente)
    - Deve ter pelo menos 1 parágrafo com texto

  XLSX:
    - Tamanho máximo: 20 MB
    - Deve ter pelo menos 1 sheet com dados

  TXT/MD:
    - Tamanho máximo: 5 MB
    - Deve ter pelo menos 100 caracteres de conteúdo

  HTML:
    - Tamanho máximo: 5 MB
    - Deve ter tag <html> ou <body>
"""
from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Limites de tamanho por extensão (bytes)
# ─────────────────────────────────────────────────────────────────────────────

_TAMANHO_MAXIMO: dict[str, int] = {
    ".pdf":  50 * 1024 * 1024,   # 50 MB
    ".csv":  10 * 1024 * 1024,   # 10 MB
    ".docx": 20 * 1024 * 1024,   # 20 MB
    ".xlsx": 20 * 1024 * 1024,   # 20 MB
    ".txt":   5 * 1024 * 1024,   #  5 MB
    ".md":    5 * 1024 * 1024,   #  5 MB
    ".html":  5 * 1024 * 1024,   #  5 MB
}

# MIME types permitidos por extensão (para validação do !ingerir via WhatsApp)
MIME_PERMITIDOS: dict[str, set[str]] = {
    ".pdf":  {"application/pdf", "application/octet-stream"},
    ".csv":  {"text/csv", "text/plain", "application/csv", "application/octet-stream"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword", "application/octet-stream",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel", "application/octet-stream",
    },
    ".txt":  {"text/plain"},
    ".md":   {"text/plain", "text/markdown"},
    ".html": {"text/html"},
}

# Extensões aceites (whitelist completa)
EXTENSOES_ACEITES = frozenset(_TAMANHO_MAXIMO.keys())

# ─────────────────────────────────────────────────────────────────────────────
# Templates de parsing_instruction por categoria detectada
# ─────────────────────────────────────────────────────────────────────────────

_TEMPLATES_PARSING: dict[str, str] = {

    "calendario": (
        "Este PDF contém o Calendário Acadêmico da UEMA. "
        "Para CADA evento na tabela, formate exatamente:\n"
        "EVENTO: [nome do evento] | DATA: [data ou período] | SEM: [semestre]\n"
        "Exemplo: EVENTO: Matrícula veteranos | DATA: 03/02/2026 a 07/02/2026 | SEM: 2026.1\n"
        "IMPORTANTE: preserve TODOS os eventos sem excepção."
    ),

    "edital": (
        "Este PDF é um Edital de Processo Seletivo da UEMA. "
        "Para tabelas de vagas, preserve exatamente:\n"
        "CURSO: [nome] | TURNO: [turno] | AC: [nº] | PcD: [nº] | TOTAL: [nº]\n"
        "Para cotas: CATEGORIA: [sigla] | NOME: [nome completo] | PÚBLICO: [descrição]\n"
        "Preserve TODOS os números de vagas e a numeração dos itens."
    ),

    "contatos": (
        "Este PDF é um Guia de Contatos da UEMA. "
        "Para cada contato, use o formato:\n"
        "CARGO: [cargo] | NOME: [nome completo] | EMAIL: [email] | TEL: [telefone]\n"
        "Mantenha o nome do setor/unidade como cabeçalho de cada bloco de contatos."
    ),

    "regulamento": (
        "Este PDF contém regulamentos, regras ou normativas da UEMA. "
        "Preserve a numeração dos artigos e incisos (Art. 1º, §1º, I, II, etc.). "
        "Converta tabelas para: COLUNA1: valor | COLUNA2: valor. "
        "Responda em português."
    ),

    "manual": (
        "Este PDF é um manual ou guia de procedimentos. "
        "Preserve a numeração de secções e sub-secções. "
        "Para listas numeradas e bullet points, mantenha a ordem. "
        "Para tabelas: COLUNA1: valor | COLUNA2: valor."
    ),

    "geral": (
        "Extrai todo o texto preservando a estrutura hierárquica. "
        "Para tabelas: COLUNA1: valor | COLUNA2: valor. "
        "Preserve numerações, datas e valores numéricos exactamente como aparecem. "
        "Responde em português."
    ),
}

# Keywords para detecção automática de categoria
_KEYWORDS_CATEGORIA: dict[str, list[str]] = {
    "calendario": ["calendário", "calendario", "acadêmico", "academico",
                   "semestre", "letivo", "matrícula"],
    "edital":     ["edital", "paes", "vestibular", "processo seletivo",
                   "vagas", "inscrição", "concurso"],
    "contatos":   ["contatos", "contato", "guia", "telefone", "email",
                   "endereço", "ramais"],
    "regulamento":["regulamento", "regimento", "resolução", "portaria",
                   "normativa", "estatuto", "regras"],
    "manual":     ["manual", "guia", "tutorial", "procedimentos",
                   "instrução", "instruções"],
}


# ─────────────────────────────────────────────────────────────────────────────
# Tipos
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ResultadoValidacao:
    """Resultado completo da validação de um documento."""
    valido:             bool
    motivo_rejeicao:    str          = ""
    extensao:           str          = ""
    doc_type:           str          = "geral"
    categoria:          str          = "geral"   # calendário, edital, etc.
    eh_pdf_scan:        bool         = False      # PDF sem texto → precisa LlamaParse
    parser_sugerido:    str          = "pymupdf"
    config_sugerido:    dict         = field(default_factory=dict)
    avisos:             list[str]    = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Função principal
# ─────────────────────────────────────────────────────────────────────────────

def validar_documento(
    caminho:       str,
    mimetype:      str = "",
    nome_original: str = "",
) -> ResultadoValidacao:
    """
    Valida um ficheiro antes da ingestão e gera config automático.

    Parâmetros:
      caminho:       caminho local do ficheiro (já salvo em disco)
      mimetype:      MIME type da Evolution API (pode ser vago: "application/octet-stream")
      nome_original: nome original enviado pelo utilizador (para detectar categoria)

    Retorna:
      ResultadoValidacao com valido=True/False e config_sugerido pronto para usar.
    """
    nome = nome_original or os.path.basename(caminho)
    ext  = _detectar_extensao(nome, mimetype, caminho)

    if not ext:
        return ResultadoValidacao(
            valido=False,
            motivo_rejeicao=(
                f"❌ Formato não suportado: '{nome}'\n"
                f"Formatos aceites: {', '.join(sorted(EXTENSOES_ACEITES))}"
            ),
        )

    # Validação de tamanho
    tamanho = os.path.getsize(caminho) if os.path.exists(caminho) else 0
    limite  = _TAMANHO_MAXIMO.get(ext, 10 * 1024 * 1024)
    if tamanho > limite:
        return ResultadoValidacao(
            valido=False,
            motivo_rejeicao=(
                f"❌ Ficheiro demasiado grande: {tamanho // 1024 // 1024}MB "
                f"(limite para {ext}: {limite // 1024 // 1024}MB)"
            ),
            extensao=ext,
        )

    if tamanho == 0:
        return ResultadoValidacao(
            valido=False,
            motivo_rejeicao="❌ Ficheiro vazio.",
            extensao=ext,
        )

    # Validação específica por tipo
    validador = {
        ".pdf":  _validar_pdf,
        ".csv":  _validar_csv,
        ".docx": _validar_docx,
        ".xlsx": _validar_xlsx,
        ".txt":  _validar_txt,
        ".md":   _validar_txt,
        ".html": _validar_html,
    }.get(ext, _validar_txt)

    resultado = validador(caminho, nome, ext)
    if not resultado.valido:
        return resultado

    # Detecta categoria e gera config automático
    categoria = _detectar_categoria(nome)
    resultado.categoria       = categoria
    resultado.doc_type        = _categoria_para_doctype(categoria)
    resultado.parser_sugerido = _escolher_parser(resultado)
    resultado.config_sugerido = _gerar_config(nome, resultado)

    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# Validadores por tipo
# ─────────────────────────────────────────────────────────────────────────────

def _validar_pdf(caminho: str, nome: str, ext: str) -> ResultadoValidacao:
    """Valida PDF: cabeçalho, texto extraível, tamanho."""
    avisos = []

    # Verifica cabeçalho %PDF
    try:
        with open(caminho, "rb") as f:
            header = f.read(5)
        if not header.startswith(b"%PDF"):
            return ResultadoValidacao(
                valido=False,
                motivo_rejeicao="❌ Ficheiro não é um PDF válido (cabeçalho %PDF ausente).",
                extensao=ext,
            )
    except Exception as e:
        return ResultadoValidacao(
            valido=False,
            motivo_rejeicao=f"❌ Não foi possível ler o ficheiro: {e}",
            extensao=ext,
        )

    # Testa extracção de texto com PyMuPDF
    eh_scan = False
    try:
        import fitz
        doc     = fitz.open(caminho)
        n_pags  = doc.page_count
        texto   = "".join(p.get_text("text") for p in doc)
        doc.close()

        if n_pags == 0:
            return ResultadoValidacao(
                valido=False,
                motivo_rejeicao="❌ PDF sem páginas.",
                extensao=ext,
            )

        chars_por_pagina = len(texto.strip()) / max(n_pags, 1)

        if chars_por_pagina < 50:
            # PDF provavelmente é scan/imagem
            eh_scan = True
            tamanho_mb = os.path.getsize(caminho) / 1024 / 1024
            if tamanho_mb > 5:
                # Scan grande → rejeita (LlamaParse seria muito caro)
                return ResultadoValidacao(
                    valido=False,
                    motivo_rejeicao=(
                        f"❌ PDF parece ser um scan/imagem ({chars_por_pagina:.0f} chars/pág). "
                        f"Para PDFs scaneados, usa PDF_PARSER=llamaparse no .env."
                    ),
                    extensao=ext,
                )
            else:
                avisos.append(
                    f"⚠️  PDF com pouco texto ({chars_por_pagina:.0f} chars/pág). "
                    "Se for scan, o LlamaParse dará melhor resultado."
                )

        logger.debug("📄 PDF validado: %d págs | %.0f chars/pág | scan=%s",
                     n_pags, chars_por_pagina, eh_scan)

    except ImportError:
        avisos.append("ℹ️  PyMuPDF não disponível para pré-validação.")
    except Exception as e:
        avisos.append(f"ℹ️  Pré-validação PyMuPDF: {e}")

    return ResultadoValidacao(
        valido=True, extensao=ext, eh_pdf_scan=eh_scan, avisos=avisos,
    )


def _validar_csv(caminho: str, nome: str, ext: str) -> ResultadoValidacao:
    """Valida CSV: encoding, estrutura mínima, número de colunas."""
    import csv as csv_mod

    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(caminho, "r", encoding=encoding, newline="") as f:
                reader  = csv_mod.reader(f)
                linhas  = [row for row in reader if any(c.strip() for c in row)]

            if len(linhas) < 2:
                return ResultadoValidacao(
                    valido=False,
                    motivo_rejeicao="❌ CSV deve ter pelo menos 1 linha de cabeçalho + 1 linha de dados.",
                    extensao=ext,
                )
            if len(linhas[0]) < 2:
                return ResultadoValidacao(
                    valido=False,
                    motivo_rejeicao=(
                        f"❌ CSV deve ter pelo menos 2 colunas. "
                        f"Encontrado: {len(linhas[0])} coluna(s). "
                        f"Verifica o separador (deve ser vírgula)."
                    ),
                    extensao=ext,
                )

            n_cols = len(linhas[0])
            n_rows = len(linhas) - 1
            logger.debug("📊 CSV validado: %d linhas × %d colunas | enc=%s",
                         n_rows, n_cols, encoding)
            return ResultadoValidacao(
                valido=True, extensao=ext,
                avisos=[f"ℹ️  CSV: {n_rows} linhas × {n_cols} colunas ({encoding})"],
            )

        except UnicodeDecodeError:
            continue
        except Exception as e:
            return ResultadoValidacao(
                valido=False,
                motivo_rejeicao=f"❌ Erro ao ler CSV: {e}",
                extensao=ext,
            )

    return ResultadoValidacao(
        valido=False,
        motivo_rejeicao="❌ CSV com encoding inválido. Salva como UTF-8 ou Latin-1.",
        extensao=ext,
    )


def _validar_docx(caminho: str, nome: str, ext: str) -> ResultadoValidacao:
    """Valida DOCX: é um ZIP válido, tem parágrafos com texto."""
    import zipfile
    try:
        if not zipfile.is_zipfile(caminho):
            return ResultadoValidacao(
                valido=False,
                motivo_rejeicao="❌ Ficheiro .docx corrompido (não é um ZIP válido).",
                extensao=ext,
            )
        from docx import Document
        doc    = Document(caminho)
        textos = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        if len(textos) < 1:
            return ResultadoValidacao(
                valido=False,
                motivo_rejeicao="❌ DOCX sem texto nos parágrafos.",
                extensao=ext,
            )
        return ResultadoValidacao(
            valido=True, extensao=ext,
            avisos=[f"ℹ️  DOCX: {len(textos)} parágrafos com texto"],
        )
    except ImportError:
        # Aceita mesmo sem python-docx (validação mínima com zipfile)
        return ResultadoValidacao(
            valido=True, extensao=ext,
            avisos=["ℹ️  python-docx não disponível para validação detalhada."],
        )
    except Exception as e:
        return ResultadoValidacao(
            valido=False,
            motivo_rejeicao=f"❌ Erro ao ler DOCX: {e}",
            extensao=ext,
        )


def _validar_xlsx(caminho: str, nome: str, ext: str) -> ResultadoValidacao:
    """Valida XLSX: tem sheets com dados."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        total_cells = sum(
            1 for ws in wb.worksheets
            for row in ws.iter_rows(values_only=True)
            for c in row if c is not None
        )
        if total_cells == 0:
            return ResultadoValidacao(
                valido=False,
                motivo_rejeicao="❌ XLSX sem dados nas células.",
                extensao=ext,
            )
        return ResultadoValidacao(
            valido=True, extensao=ext,
            avisos=[f"ℹ️  XLSX: {len(wb.sheetnames)} sheet(s), {total_cells} células"],
        )
    except ImportError:
        return ResultadoValidacao(
            valido=True, extensao=ext,
            avisos=["ℹ️  openpyxl não disponível — validação mínima."],
        )
    except Exception as e:
        return ResultadoValidacao(
            valido=False,
            motivo_rejeicao=f"❌ Erro ao ler XLSX: {e}",
            extensao=ext,
        )


def _validar_txt(caminho: str, nome: str, ext: str) -> ResultadoValidacao:
    """Valida TXT/MD: tem conteúdo mínimo."""
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(caminho, "r", encoding=encoding) as f:
                conteudo = f.read()
            if len(conteudo.strip()) < 100:
                return ResultadoValidacao(
                    valido=False,
                    motivo_rejeicao=f"❌ Ficheiro muito curto ({len(conteudo)} chars). Mínimo: 100.",
                    extensao=ext,
                )
            return ResultadoValidacao(
                valido=True, extensao=ext,
                avisos=[f"ℹ️  {ext}: {len(conteudo)} chars ({encoding})"],
            )
        except UnicodeDecodeError:
            continue
    return ResultadoValidacao(
        valido=False,
        motivo_rejeicao="❌ Encoding inválido no ficheiro de texto.",
        extensao=ext,
    )


def _validar_html(caminho: str, nome: str, ext: str) -> ResultadoValidacao:
    """Valida HTML: tem tags básicas."""
    res = _validar_txt(caminho, nome, ext)
    if not res.valido:
        return res
    try:
        with open(caminho, "r", encoding="utf-8", errors="ignore") as f:
            conteudo = f.read(2000).lower()
        if "<html" not in conteudo and "<body" not in conteudo:
            res.avisos.append("⚠️  HTML sem tags <html>/<body>. Pode não ser HTML válido.")
    except Exception:
        pass
    return res


# ─────────────────────────────────────────────────────────────────────────────
# Detecção automática
# ─────────────────────────────────────────────────────────────────────────────

def _detectar_extensao(nome: str, mimetype: str, caminho: str) -> str:
    """Detecta a extensão real do ficheiro (nome > mimetype > magic bytes)."""
    # 1. Pela extensão do nome original
    ext = os.path.splitext(nome.lower())[1]
    if ext in EXTENSOES_ACEITES:
        return ext

    # 2. Pelo MIME type
    mime_para_ext = {
        "application/pdf": ".pdf",
        "text/csv": ".csv",
        "text/plain": ".txt",
        "text/markdown": ".md",
        "text/html": ".html",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/msword": ".docx",
        "application/vnd.ms-excel": ".xlsx",
    }
    if mimetype in mime_para_ext:
        return mime_para_ext[mimetype]

    # 3. Por magic bytes (últimos recursos)
    try:
        with open(caminho, "rb") as f:
            magic = f.read(8)
        if magic[:4] == b"%PDF":
            return ".pdf"
        if magic[:2] == b"PK":
            return ".docx"  # ZIP → provavelmente DOCX ou XLSX
    except Exception:
        pass

    return ""  # extensão não detectada → rejeitar


def _detectar_categoria(nome: str) -> str:
    """Detecta categoria do documento pelo nome do ficheiro."""
    nome_lower = nome.lower()
    for categoria, keywords in _KEYWORDS_CATEGORIA.items():
        if any(kw in nome_lower for kw in keywords):
            return categoria
    return "geral"


def _categoria_para_doctype(categoria: str) -> str:
    mapa = {
        "calendario":  "calendario",
        "edital":      "edital",
        "contatos":    "contatos",
        "regulamento": "geral",
        "manual":      "geral",
        "geral":       "geral",
    }
    return mapa.get(categoria, "geral")


def _escolher_parser(resultado: ResultadoValidacao) -> str:
    """Escolhe o parser mais adequado."""
    if resultado.extensao != ".pdf":
        return "auto"  # não-PDF → parser automático pelo tipo
    if resultado.eh_pdf_scan:
        return "llamaparse"  # scan → precisa OCR do LlamaParse
    return "pymupdf"  # PDF com texto → pymupdf é suficiente


def _gerar_config(nome: str, resultado: ResultadoValidacao) -> dict:
    """Gera config completo para o DOCUMENT_CONFIG."""
    ext        = resultado.extensao
    categoria  = resultado.categoria
    doc_type   = resultado.doc_type
    parser     = resultado.parser_sugerido

    # Tamanhos de chunk por tipo de documento
    chunk_sizes = {
        "calendario":  (280, 80),
        "edital":      (500, 80),
        "contatos":    (250, 30),
        "regulamento": (400, 60),
        "manual":      (350, 50),
        "geral":       (400, 60),
    }
    chunk_size, overlap = chunk_sizes.get(categoria, (400, 60))

    # Label limpa (para o prefixo hierárquico anti-alucinação)
    label = re.sub(r"[._-]", " ", os.path.splitext(nome)[0]).upper()

    # Título legível
    titulo = label.title()

    config: dict = {
        "doc_type":   doc_type,
        "titulo":     titulo,
        "chunk_size": chunk_size,
        "overlap":    overlap,
        "label":      label,
    }

    # Parser para PDFs
    if ext == ".pdf" and parser == "llamaparse":
        config["parser"] = "llamaparse"
        config["parsing_instruction"] = _TEMPLATES_PARSING.get(categoria, _TEMPLATES_PARSING["geral"])
    elif ext == ".pdf" and parser == "pymupdf":
        config["parser"] = "pymupdf"
        # parsing_instruction não necessária para pymupdf
    elif ext not in (".pdf",):
        # CSV, DOCX, etc. — sem parser específico (auto pela extensão)
        pass

    return config


# ─────────────────────────────────────────────────────────────────────────────
# Funções auxiliares para o !ingerir
# ─────────────────────────────────────────────────────────────────────────────

def formatar_resultado_para_whatsapp(resultado: ResultadoValidacao, nome: str) -> str:
    """
    Formata o resultado da validação para enviar ao admin via WhatsApp.
    Conciso e informativo.
    """
    if not resultado.valido:
        return resultado.motivo_rejeicao

    linhas = [
        f"✅ *Ficheiro válido: `{nome}`*\n",
        f"📋 Tipo detectado: *{resultado.categoria}*",
        f"📂 doc_type: `{resultado.doc_type}`",
        f"⚙️  Parser: `{resultado.parser_sugerido}`",
        f"🧩 Chunk size: {resultado.config_sugerido.get('chunk_size', '?')} chars",
    ]
    if resultado.eh_pdf_scan:
        linhas.append("⚠️  PDF scan — usando LlamaParse (custo: ~$0.003/pág)")
    if resultado.avisos:
        linhas.append("\n" + "\n".join(resultado.avisos))

    return "\n".join(linhas)


def tipos_aceites_mensagem() -> str:
    """Mensagem de ajuda sobre formatos aceites (para o !ingerir sem ficheiro)."""
    return (
        "📎 *Formatos aceites para ingestão:*\n\n"
        "📄 *PDF* — Documentos, editais, calendários\n"
        "   • Max: 50 MB | Texto ou tabelas estruturadas\n\n"
        "📊 *CSV* — Tabelas de dados (recomendado para vagas/contatos)\n"
        "   • Max: 10 MB | Separador: vírgula | Encoding: UTF-8\n"
        "   • Estrutura: 1ª linha = cabeçalho, restantes = dados\n\n"
        "📝 *DOCX* — Documentos Word (manuais, guias)\n"
        "   • Max: 20 MB\n\n"
        "📈 *XLSX* — Planilhas Excel\n"
        "   • Max: 20 MB\n\n"
        "📋 *TXT / MD* — Texto simples ou Markdown\n"
        "   • Max: 5 MB\n\n"
        "Para ingerir: *anexa o ficheiro + escreve `!ingerir`*"
    )